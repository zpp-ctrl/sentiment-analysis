# -*- coding: utf-8 -*-
"""快速生成情感分析Excel报告"""
import sys
sys.path.insert(0, '.')
import pandas as pd
from datetime import datetime
from config import OUTPUT_DIR
from module2_text_sentiment import classify_sentiment
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

print('>>> Loading data...')

# Load resume data (all accumulated posts)
df = pd.read_csv('output/douyin_posts_resume.csv')
total = len(df)
print(f'    Loaded {total} posts')

# Run sentiment analysis on all posts
print('>>> Running sentiment analysis...')
sentiments = []
scores = []
for i, text in enumerate(df['post_content'].fillna('')):
    label, score = classify_sentiment(str(text))
    sentiments.append(label)
    scores.append(score)
    if (i + 1) % 500 == 0:
        print(f'    Processed {i+1}/{total}...')

df['sentiment'] = sentiments
df['sentiment_score'] = scores

# Save to CSV with sentiment
csv_path = os.path.join(OUTPUT_DIR, f'{datetime.now().strftime("%Y%m%d")}_posts_with_sentiment.csv')
df.to_csv(csv_path, index=False, encoding='utf-8-sig')
print(f'    CSV saved: {csv_path} ({len(df)} rows)')

# Compute stats
bullish_cnt = sum(1 for s in sentiments if s == 'bullish')
bearish_cnt = sum(1 for s in sentiments if s == 'bearish')
neutral_cnt = sum(1 for s in sentiments if s == 'neutral')
avg_score = sum(scores) / len(scores) if scores else 0
ratio = bullish_cnt / bearish_cnt if bearish_cnt > 0 else float('inf')

print(f'    Bullish={bullish_cnt}, Bearish={bearish_cnt}, Neutral={neutral_cnt}, Ratio={ratio:.2f}')

# ========================================
# Build Excel Report
# ========================================
print('>>> Building Excel report...')

wb = Workbook()

# --- Styles ---
title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
section_font = Font(name='Arial', size=13, bold=True)
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
data_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=11, bold=True)
small_font = Font(name='Arial', size=9, italic=True, color='666666')

title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
no_fill = PatternFill(fill_type=None)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ========================================
# Sheet 1: Sentiment Overview
# ========================================
ws = wb.active
ws.title = 'Sentiment Overview'

ws.merge_cells('A1:F1')
ws['A1'] = f'Daily Financial Sentiment Report - {datetime.now().strftime("%Y-%m-%d")}'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = center
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:F2')
ws['A2'] = f'Source: {total} Douyin finance posts | Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
ws['A2'].font = small_font
ws['A2'].alignment = center

# --- Section 1: Key Metrics ---
row = 4
ws.merge_cells(f'A{row}:F{row}')
ws[f'A{row}'] = '1. Key Sentiment Metrics'
ws[f'A{row}'].font = section_font

row = 6
headers = ['Metric', 'Value', 'Description']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

if ratio > 2.0:
    bias = 'Strongly Bullish'
elif ratio > 1.5:
    bias = 'Bullish'
elif ratio > 1.1:
    bias = 'Slightly Bullish'
elif ratio > 0.9:
    bias = 'Neutral'
elif ratio > 0.5:
    bias = 'Slightly Bearish'
else:
    bias = 'Bearish'

metrics = [
    ('Total Posts', total, 'All financial posts collected'),
    ('Bullish Posts', bullish_cnt, 'Posts with optimistic market outlook'),
    ('Bearish Posts', bearish_cnt, 'Posts with pessimistic market outlook'),
    ('Neutral Posts', neutral_cnt, 'Posts with balanced/neutral views'),
    ('Bullish/Bearish Ratio', f'{ratio:.2f} : 1', f'Market Bias: {bias}'),
    ('Bullish Percentage', f'{bullish_cnt/total*100:.1f}%', ''),
    ('Bearish Percentage', f'{bearish_cnt/total*100:.1f}%', ''),
    ('Avg Sentiment Score', f'{avg_score:.3f}', '0=Extreme Bearish, 0.5=Neutral, 1=Extreme Bullish'),
]

for i, (label, val, desc) in enumerate(metrics):
    r = row + 1 + i
    c1 = ws.cell(row=r, column=1, value=label)
    c1.font = bold_font
    c1.border = thin_border
    c1.fill = light_fill if i % 2 == 0 else no_fill

    c2 = ws.cell(row=r, column=2, value=val)
    c2.font = bold_font
    c2.alignment = center
    c2.border = thin_border
    c2.fill = light_fill if i % 2 == 0 else no_fill

    if 'Ratio' in label:
        c2.fill = green_fill if ratio > 1 else (red_fill if ratio < 1 else yellow_fill)

    c3 = ws.cell(row=r, column=3, value=desc)
    c3.font = data_font
    c3.border = thin_border
    c3.fill = light_fill if i % 2 == 0 else no_fill

# --- Section 2: Distribution ---
row2 = r + 2
ws.merge_cells(f'A{row2}:F{row2}')
ws[f'A{row2}'] = '2. Sentiment Distribution'
ws[f'A{row2}'].font = section_font

row2 += 2
for i, h in enumerate(['Sentiment', 'Count', 'Percentage', 'Visualization'], 1):
    c = ws.cell(row=row2, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

dist = [
    ('Bullish (Long)', bullish_cnt, bullish_cnt/total*100, green_fill),
    ('Bearish (Short)', bearish_cnt, bearish_cnt/total*100, red_fill),
    ('Neutral', neutral_cnt, neutral_cnt/total*100, yellow_fill),
]
for i, (label, cnt, pct, color) in enumerate(dist):
    r = row2 + 1 + i
    c1 = ws.cell(row=r, column=1, value=label)
    c1.font = bold_font
    c1.fill = color
    c1.border = thin_border

    c2 = ws.cell(row=r, column=2, value=cnt)
    c2.font = data_font
    c2.alignment = center
    c2.border = thin_border

    c3 = ws.cell(row=r, column=3, value=f'{pct:.1f}%')
    c3.font = data_font
    c3.alignment = center
    c3.border = thin_border

    bar_len = int(pct)
    bar = '|' + '#' * bar_len + '|'
    c4 = ws.cell(row=r, column=4, value=bar)
    c4.font = Font(name='Consolas', size=10)
    c4.border = thin_border

# --- Section 3: Daily Trend ---
row3 = r + 2
ws.merge_cells(f'A{row3}:F{row3}')
ws[f'A{row3}'] = '3. Daily Sentiment Trend'
ws[f'A{row3}'].font = section_font

row3 += 2
headers3 = ['Date', 'Total Posts', 'Bullish', 'Bearish', 'Ratio', 'Bias']
for i, h in enumerate(headers3, 1):
    c = ws.cell(row=row3, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

df['date'] = pd.to_datetime(df['collect_time']).dt.date
daily_stats = df.groupby('date').agg(
    posts=('post_id', 'count'),
    bullish=('sentiment', lambda x: sum(1 for v in x if v == 'bullish')),
    bearish=('sentiment', lambda x: sum(1 for v in x if v == 'bearish')),
).reset_index().sort_values('date')

for j, (_, drow) in enumerate(daily_stats.tail(14).iterrows()):
    r = row3 + 1 + j
    d = drow['date']
    b = int(drow['bullish'])
    s = int(drow['bearish'])
    dr = b / s if s > 0 else float('inf')
    if dr > 1.5:
        trend = 'Bullish'
    elif dr > 1.1:
        trend = 'Slightly Bullish'
    elif dr > 0.9:
        trend = 'Neutral'
    elif dr > 0.5:
        trend = 'Slightly Bearish'
    else:
        trend = 'Bearish'

    fill = light_fill if j % 2 == 0 else no_fill
    for ci, v in enumerate([str(d), int(drow['posts']), b, s, f'{dr:.2f}', trend], 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.font = data_font
        c.alignment = center
        c.border = thin_border
        if fill:
            c.fill = fill
        if ci == 5:
            c.fill = green_fill if dr > 1.1 else (red_fill if dr < 0.9 else yellow_fill)
        if ci == 6:
            c.fill = green_fill if 'Bull' in trend else (red_fill if 'Bear' in trend else yellow_fill)

# ========================================
# Sheet 2: Top Accounts
# ========================================
ws2 = wb.create_sheet('Top Accounts')

ws2.merge_cells('A1:H1')
ws2['A1'] = f'Top Financial Bloggers ({datetime.now().strftime("%Y-%m-%d")})'
ws2['A1'].font = title_font
ws2['A1'].fill = title_fill
ws2['A1'].alignment = center
ws2.row_dimensions[1].height = 35

row = 3
headers2 = ['Rank', 'Account Name', 'Total Posts', 'Bullish', 'Bearish', 'Bull/Bear Ratio', 'Avg Score', 'Bias']
for i, h in enumerate(headers2, 1):
    c = ws2.cell(row=row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

top = df.groupby('account_name').agg(
    posts=('post_id', 'count'),
    bullish=('sentiment', lambda x: sum(1 for v in x if v == 'bullish')),
    bearish=('sentiment', lambda x: sum(1 for v in x if v == 'bearish')),
    avg_score=('sentiment_score', 'mean'),
).sort_values('posts', ascending=False).head(30)

for rank, (name, rd) in enumerate(top.iterrows(), 1):
    r = row + rank
    bs = int(rd['bullish'])
    br = int(rd['bearish'])
    ar = bs / br if br > 0 else float('inf')
    if ar > 1.5:
        bias = 'Bullish'
    elif ar > 1.1:
        bias = 'Slightly Bullish'
    elif ar > 0.9:
        bias = 'Neutral'
    elif ar > 0.5:
        bias = 'Slightly Bearish'
    else:
        bias = 'Bearish'

    vals = [rank, str(name)[:35], int(rd['posts']), bs, br, f'{ar:.2f}', round(rd['avg_score'], 3), bias]
    for i, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=i, value=v)
        c.font = data_font
        c.alignment = center
        c.border = thin_border
        if i == 6:
            c.fill = green_fill if ar > 1.1 else (red_fill if ar < 0.9 else yellow_fill)
        if i == 8:
            c.fill = green_fill if 'Bull' in bias else (red_fill if 'Bear' in bias else yellow_fill)

# ========================================
# Sheet 3: All Posts
# ========================================
ws3 = wb.create_sheet('All Posts')

ws3.merge_cells('A1:H1')
ws3['A1'] = 'All Posts with Sentiment Analysis'
ws3['A1'].font = title_font
ws3['A1'].fill = title_fill
ws3['A1'].alignment = center
ws3.row_dimensions[1].height = 35

row = 3
headers3 = ['#', 'Account', 'Content', 'Type', 'Likes', 'Post Time', 'Sentiment', 'Score']
for i, h in enumerate(headers3, 1):
    c = ws3.cell(row=row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

# Show up to 500 posts to keep file manageable
for idx, (_, post) in enumerate(df.head(500).iterrows(), 1):
    r = row + idx
    s = str(post.get('sentiment', 'neutral'))
    vals = [
        idx,
        str(post.get('account_name', ''))[:30],
        str(post.get('post_content', ''))[:120],
        str(post.get('post_type', ''))[:10],
        post.get('like_count', 0) if pd.notna(post.get('like_count')) else 0,
        str(post.get('post_time', ''))[:19],
        s,
        round(post.get('sentiment_score', 0.5), 3) if pd.notna(post.get('sentiment_score')) else 0.5,
    ]
    for i, v in enumerate(vals, 1):
        c = ws3.cell(row=r, column=i, value=v)
        c.font = data_font
        c.border = thin_border
        if i in (1, 4, 5, 8):
            c.alignment = center
        elif i == 7:
            c.alignment = center
            c.fill = green_fill if s == 'bullish' else (red_fill if s == 'bearish' else yellow_fill)
            c.font = bold_font

# Column widths
for sht_name, col_widths in [
    ('Sentiment Overview', {'A': 28, 'B': 20, 'C': 55, 'D': 40, 'E': 12, 'F': 12}),
    ('Top Accounts', {'A': 8, 'B': 38, 'C': 14, 'D': 10, 'E': 10, 'F': 14, 'G': 12, 'H': 18}),
    ('All Posts', {'A': 6, 'B': 30, 'C': 70, 'D': 10, 'E': 10, 'F': 20, 'G': 14, 'H': 8}),
]:
    sht = wb[sht_name]
    for col, w in col_widths.items():
        sht.column_dimensions[col].width = w

# Freeze panes
ws.freeze_panes = 'A7'
ws2.freeze_panes = 'A4'
ws3.freeze_panes = 'A4'

# Auto-filter
ws.auto_filter.ref = f'A{row3}:F{row3}'
ws2.auto_filter.ref = f'A{row}:H{row}'

# Save
excel_path = os.path.join(OUTPUT_DIR, f'{datetime.now().strftime("%Y%m%d")}_Finance_Sentiment_Report.xlsx')
wb.save(excel_path)

print(f'>>> Report saved: {excel_path}')
print(f'>>> CSV saved: {csv_path}')
print('=' * 60)
print(f'    Total Posts: {total}')
print(f'    Bullish:  {bullish_cnt:>5} ({bullish_cnt/total*100:5.1f}%)')
print(f'    Bearish:  {bearish_cnt:>5} ({bearish_cnt/total*100:5.1f}%)')
print(f'    Neutral:  {neutral_cnt:>5} ({neutral_cnt/total*100:5.1f}%)')
print(f'    Ratio:    {ratio:.2f} : 1')
print(f'    Avg Score:{avg_score:.3f}')
print(f'    Market:   {bias}')
print('=' * 60)
print('DONE')
