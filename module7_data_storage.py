# -*- coding: utf-8 -*-
"""
============================================================
模块7: MySQL数据读写 & Excel报表导出
============================================================
功能:
  1. MySQL连接管理、自动重连
  2. 原始帖子表增量写入
  3. 每日情绪统计表增量写入
  4. 预测回测记录表增量写入/回填更新
  5. Excel日报自动导出，格式【日期_财经舆情预测日报.xlsx】
"""

import os
import re
import time
import logging
from datetime import date, datetime
from typing import Dict, List, Optional

import pandas as pd
import pymysql
from pymysql.cursors import DictCursor
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter

from config import (
    MYSQL_CONFIG,
    OUTPUT_DIR,
    EXCEL_REPORT_TEMPLATE,
    MAX_RETRY_COUNT,
    RETRY_DELAY_SECONDS,
    INDEX_CODE_MAP,
)

logger = logging.getLogger(__name__)

# ★ v7: SQLite回退数据库路径
SQLITE_DB_PATH = os.path.join(OUTPUT_DIR, "financial_sentiment.db")


def _safe_int(value, default=0):
    """安全转 int，容忍 None/NaN/空值，避免 'cannot convert float NaN to integer'"""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value, default=0.0):
    """安全转 float，容忍 None/NaN/空值"""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


class MySQLManager:
    """
    MySQL数据库管理器
    提供连接管理、表写入、查询等基础操作
    """

    def __init__(self, config: Optional[Dict] = None):
        self.config = config or MYSQL_CONFIG
        self.connection = None
        self._mysql_available = None  # None=未检测, True/False

    def is_mysql_available(self) -> bool:
        """检测MySQL是否可用"""
        if self._mysql_available is not None:
            return self._mysql_available
        try:
            import socket
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            result = sock.connect_ex(('127.0.0.1', 3306))
            sock.close()
            self._mysql_available = (result == 0)
        except Exception:
            self._mysql_available = False
        return self._mysql_available

    def connect(self):
        """建立数据库连接（含重试），MySQL不可用时自动回退到SQLite"""
        # 先尝试 MySQL
        for attempt in range(1, MAX_RETRY_COUNT + 1):
            try:
                self.connection = pymysql.connect(**self.config)
                self._backend = "mysql"
                logger.info("MySQL连接成功: %s:%d/%s",
                            self.config["host"], self.config["port"],
                            self.config["database"])
                self._init_mysql_tables()
                return
            except Exception as exc:
                logger.warning("MySQL连接失败 [尝试 %d/%d]: %s",
                               attempt, MAX_RETRY_COUNT, str(exc))
                if attempt < MAX_RETRY_COUNT:
                    time.sleep(RETRY_DELAY_SECONDS * attempt)

        # ★ v7: SQLite回退
        logger.warning("=" * 60)
        logger.warning("MySQL不可用，自动切换到SQLite回退模式")
        logger.warning("SQLite数据库: %s", SQLITE_DB_PATH)
        logger.warning("数据保存在本地文件，MySQL恢复后可迁移")
        logger.warning("=" * 60)
        try:
            import sqlite3
            self.connection = sqlite3.connect(SQLITE_DB_PATH, check_same_thread=False)
            self._backend = "sqlite"
            self._init_sqlite_tables()
            logger.info("SQLite连接成功: %s", SQLITE_DB_PATH)
            return
        except Exception as exc:
            raise ConnectionError(f"数据库连接失败(MySQL+SQLite均不可用): {exc}")

    def close(self):
        """关闭连接"""
        if self.connection:
            try:
                self.connection.close()
            except Exception:
                pass
            self.connection = None
            self._backend = None

    def ensure_connection(self):
        """确保连接可用，断线自动重连"""
        if self.connection is None:
            self.connect()
            return
        backend = getattr(self, '_backend', 'mysql')
        if backend == 'mysql':
            try:
                self.connection.ping(reconnect=True)
            except Exception:
                self.connect()

    def _init_mysql_tables(self):
        """MySQL建表（使用init_db.sql的内容）"""
        sql = """
        CREATE TABLE IF NOT EXISTS raw_posts (
            id BIGINT AUTO_INCREMENT PRIMARY KEY,
            post_id VARCHAR(128) NOT NULL, platform VARCHAR(32) NOT NULL,
            account_id VARCHAR(64) NOT NULL, account_name VARCHAR(256) NOT NULL,
            account_level VARCHAR(32) NOT NULL, post_type VARCHAR(32) NOT NULL,
            post_content TEXT NOT NULL, post_url VARCHAR(512) DEFAULT NULL,
            like_count INT DEFAULT 0, comment_count INT DEFAULT 0,
            share_count INT DEFAULT 0, post_time DATETIME NOT NULL,
            collect_time DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            is_ad TINYINT(1) DEFAULT 0, is_stock_only TINYINT(1) DEFAULT 0,
            sentiment VARCHAR(16) DEFAULT NULL, sentiment_score DECIMAL(5,4) DEFAULT NULL,
            video_content_extracted TINYINT(1) DEFAULT 0 COMMENT '是否已提取视频内容: 0=否, 1=是',
            audio_transcript TEXT DEFAULT NULL COMMENT '视频语音转录文本',
            visual_description TEXT DEFAULT NULL COMMENT '视频画面分析描述',
            UNIQUE KEY uk_post (post_id, platform)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        CREATE TABLE IF NOT EXISTS daily_sentiment_stats (
            id BIGINT AUTO_INCREMENT PRIMARY KEY,
            stat_date DATE NOT NULL, account_level VARCHAR(32) NOT NULL,
            total_posts INT DEFAULT 0, bullish_count INT DEFAULT 0,
            bearish_count INT DEFAULT 0, neutral_count INT DEFAULT 0,
            bullish_ratio DECIMAL(6,4) DEFAULT 0, bearish_ratio DECIMAL(6,4) DEFAULT 0,
            neutral_ratio DECIMAL(6,4) DEFAULT 0, avg_sentiment_score DECIMAL(5,4) DEFAULT 0,
            create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uk_date_level (stat_date, account_level)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        CREATE TABLE IF NOT EXISTS prediction_backtest_records (
            id BIGINT AUTO_INCREMENT PRIMARY KEY,
            predict_date DATE NOT NULL, target_date DATE NOT NULL,
            index_code VARCHAR(32) NOT NULL, index_name VARCHAR(64) NOT NULL,
            predict_direction VARCHAR(16) NOT NULL, predict_prob DECIMAL(6,4) NOT NULL,
            actual_direction VARCHAR(16) DEFAULT NULL,
            actual_pct_change DECIMAL(8,4) DEFAULT NULL,
            is_correct TINYINT(1) DEFAULT NULL, backtest_time DATETIME DEFAULT NULL,
            create_time DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uk_predict (predict_date, index_code)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """
        try:
            for stmt in sql.split(";"):
                stmt = stmt.strip()
                if stmt:
                    with self.connection.cursor() as c:
                        c.execute(stmt)
            self.connection.commit()
            logger.info("MySQL建表完成")
        except Exception as e:
            logger.debug("MySQL建表(可能已存在): %s", str(e)[:60])

    def _init_sqlite_tables(self):
        """SQLite建表（兼容MySQL语法）"""
        import sqlite3
        sql = """
        CREATE TABLE IF NOT EXISTS raw_posts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            post_id TEXT NOT NULL, platform TEXT NOT NULL,
            account_id TEXT NOT NULL, account_name TEXT NOT NULL,
            account_level TEXT NOT NULL, post_type TEXT NOT NULL,
            post_content TEXT NOT NULL, post_url TEXT,
            like_count INTEGER DEFAULT 0, comment_count INTEGER DEFAULT 0,
            share_count INTEGER DEFAULT 0, post_time TEXT NOT NULL,
            collect_time TEXT NOT NULL,
            is_ad INTEGER DEFAULT 0, is_stock_only INTEGER DEFAULT 0,
            sentiment TEXT, sentiment_score REAL DEFAULT NULL,
            video_content_extracted INTEGER DEFAULT 0,
            audio_transcript TEXT,
            visual_description TEXT,
            UNIQUE(post_id, platform)
        );
        CREATE TABLE IF NOT EXISTS daily_sentiment_stats (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            stat_date TEXT NOT NULL, account_level TEXT NOT NULL,
            total_posts INTEGER DEFAULT 0, bullish_count INTEGER DEFAULT 0,
            bearish_count INTEGER DEFAULT 0, neutral_count INTEGER DEFAULT 0,
            bullish_ratio REAL DEFAULT 0, bearish_ratio REAL DEFAULT 0,
            neutral_ratio REAL DEFAULT 0, avg_sentiment_score REAL DEFAULT 0,
            create_time TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(stat_date, account_level)
        );
        CREATE TABLE IF NOT EXISTS prediction_backtest_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            predict_date TEXT NOT NULL, target_date TEXT NOT NULL,
            index_code TEXT NOT NULL, index_name TEXT NOT NULL,
            predict_direction TEXT NOT NULL, predict_prob REAL NOT NULL,
            actual_direction TEXT, actual_pct_change REAL,
            is_correct INTEGER, backtest_time TEXT,
            create_time TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(predict_date, index_code)
        );
        """
        try:
            for stmt in sql.split(";"):
                stmt = stmt.strip()
                if stmt:
                    self.connection.execute(stmt)
            self.connection.commit()
            logger.info("SQLite建表完成")
        except Exception as e:
            logger.debug("SQLite建表(可能已存在): %s", str(e)[:60])

    def _convert_mysql_to_sqlite(self, sql: str) -> str:
        """将MySQL特有语法转换为SQLite兼容语法"""
        # 1. INSERT IGNORE → INSERT OR IGNORE
        sql = sql.replace('INSERT IGNORE INTO', 'INSERT OR IGNORE INTO')
        # 2. ON DUPLICATE KEY UPDATE → INSERT OR REPLACE
        if re.search(r'\bON\s+DUPLICATE\s+KEY\s+UPDATE\b', sql, re.IGNORECASE):
            # 剥离 ON DUPLICATE KEY UPDATE 及其后面的所有内容
            sql = re.sub(r'\s+ON\s+DUPLICATE\s+KEY\s+UPDATE\s+.*$', '', sql, flags=re.IGNORECASE | re.DOTALL)
            # 将 INSERT INTO 替换为 INSERT OR REPLACE INTO
            sql = re.sub(r'INSERT\s+INTO', 'INSERT OR REPLACE INTO', sql, count=1, flags=re.IGNORECASE)
        # 3. %s → ? (MySQL→SQLite占位符)
        sql = sql.replace('%s', '?')
        return sql

    def execute(self, sql: str, params: Optional[tuple] = None):
        """执行SQL语句（兼容MySQL和SQLite）"""
        self.ensure_connection()
        backend = getattr(self, '_backend', 'mysql')
        try:
            if backend == 'sqlite':
                sql = self._convert_mysql_to_sqlite(sql)
                if params:
                    self.connection.execute(sql, params)
                else:
                    self.connection.execute(sql)
                self.connection.commit()
            else:
                with self.connection.cursor() as cursor:
                    cursor.execute(sql, params)
                self.connection.commit()
        except Exception as exc:
            try:
                if backend == 'mysql':
                    self.connection.rollback()
            except Exception:
                pass
            logger.error("SQL执行失败: %s | SQL: %s", str(exc), sql[:200])
            raise

    def execute_many(self, sql: str, params_list: List[tuple]):
        """批量执行SQL语句（兼容MySQL和SQLite）"""
        self.ensure_connection()
        backend = getattr(self, '_backend', 'mysql')
        try:
            if backend == 'sqlite':
                sql = self._convert_mysql_to_sqlite(sql)
                self.connection.executemany(sql, params_list)
                self.connection.commit()
            else:
                with self.connection.cursor() as cursor:
                    cursor.executemany(sql, params_list)
                self.connection.commit()
        except Exception as exc:
            try:
                if backend == 'mysql':
                    self.connection.rollback()
            except Exception:
                pass
            logger.error("批量SQL执行失败: %s", str(exc))
            raise

    def query(self, sql: str, params: Optional[tuple] = None) -> pd.DataFrame:
        """执行查询并返回DataFrame"""
        self.ensure_connection()
        backend = getattr(self, '_backend', 'mysql')
        try:
            if backend == 'sqlite':
                sql = self._convert_mysql_to_sqlite(sql)
                return pd.read_sql_query(sql, self.connection, params=params)
            return pd.read_sql(sql, self.connection, params=params)
        except Exception as exc:
            logger.error("SQL查询失败: %s", str(exc))
            return pd.DataFrame()

    # ==================== 表写入方法 ====================

    def insert_raw_posts(self, df_posts: pd.DataFrame) -> int:
        """
        批量插入原始帖子（增量追加，已存在则跳过）
        返回实际插入条数
        """
        if df_posts is None or len(df_posts) == 0:
            logger.warning("无帖子数据需要写入")
            return 0

        insert_sql = """
            INSERT IGNORE INTO raw_posts
                (post_id, platform, account_id, account_name, account_level,
                 post_type, post_content, post_url, like_count, comment_count,
                 share_count, post_time, collect_time, is_ad, is_stock_only,
                 sentiment, sentiment_score,
                 video_content_extracted, audio_transcript, visual_description)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        records = []
        for _, row in df_posts.iterrows():
            records.append((
                str(row.get("post_id", "")),
                str(row.get("platform", "")),
                str(row.get("account_id", "")),
                str(row.get("account_name", "")),
                str(row.get("account_level", "")),
                str(row.get("post_type", "")),
                str(row.get("post_content", "")),
                str(row.get("post_url", "")) if pd.notna(row.get("post_url")) else None,
                _safe_int(row.get("like_count", 0)),
                _safe_int(row.get("comment_count", 0)),
                _safe_int(row.get("share_count", 0)),
                row.get("post_time", datetime.now()),
                row.get("collect_time", datetime.now()),
                _safe_int(row.get("is_ad", 0)),
                _safe_int(row.get("is_stock_only", 0)),
                str(row.get("sentiment", "")),
                _safe_float(row.get("sentiment_score", 0.5)),
                # ★ 视频内容提取字段
                _safe_int(row.get("video_content_extracted", False)),
                str(row.get("audio_transcript", "")) if pd.notna(row.get("audio_transcript")) else None,
                str(row.get("visual_description", "")) if pd.notna(row.get("visual_description")) else None,
            ))

        try:
            self.execute_many(insert_sql, records)
            logger.info("原始帖子写入完成: 提交 %d 条", len(records))
            return len(records)
        except Exception as exc:
            logger.error("原始帖子写入失败: %s", str(exc))
            return 0

    def insert_sentiment_stats(self, stats_df: pd.DataFrame) -> int:
        """
        写入每日情绪统计（UNIQUE KEY: stat_date + account_level）
        已存在则更新
        """
        if stats_df is None or len(stats_df) == 0:
            return 0

        upsert_sql = """
            INSERT INTO daily_sentiment_stats
                (stat_date, account_level, total_posts,
                 bullish_count, bearish_count, neutral_count,
                 bullish_ratio, bearish_ratio, neutral_ratio,
                 avg_sentiment_score)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                total_posts = VALUES(total_posts),
                bullish_count = VALUES(bullish_count),
                bearish_count = VALUES(bearish_count),
                neutral_count = VALUES(neutral_count),
                bullish_ratio = VALUES(bullish_ratio),
                bearish_ratio = VALUES(bearish_ratio),
                neutral_ratio = VALUES(neutral_ratio),
                avg_sentiment_score = VALUES(avg_sentiment_score)
        """

        records = []
        for _, row in stats_df.iterrows():
            records.append((
                row.get("stat_date"),
                str(row.get("account_level", "")),
                int(row.get("total_posts", 0)),
                int(row.get("bullish_count", 0)),
                int(row.get("bearish_count", 0)),
                int(row.get("neutral_count", 0)),
                float(row.get("bullish_ratio", 0.0)),
                float(row.get("bearish_ratio", 0.0)),
                float(row.get("neutral_ratio", 0.0)),
                float(row.get("avg_sentiment_score", 0.5)),
            ))

        try:
            self.execute_many(upsert_sql, records)
            logger.info("情绪统计写入完成: %d 条, stat_date=%s",
                        len(records), stats_df.iloc[0].get("stat_date", ""))
            return len(records)
        except Exception as exc:
            logger.error("情绪统计写入失败: %s", str(exc))
            return 0

    def insert_predictions(self, predictions: List[Dict]) -> int:
        """
        写入预测记录（UNIQUE KEY: predict_date + index_code）
        基于情感分析预测四大指数次日涨跌方向
        """
        if not predictions:
            return 0

        insert_sql = """
            INSERT INTO prediction_backtest_records
                (predict_date, target_date, index_code, index_name,
                 predict_direction, predict_prob)
            VALUES
                (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                predict_direction = VALUES(predict_direction),
                predict_prob = VALUES(predict_prob)
        """

        records = []
        for pred in predictions:
            records.append((
                pred.get("predict_date"),
                pred.get("target_date"),
                pred.get("index_code"),
                pred.get("index_name"),
                pred.get("predict_direction"),
                pred.get("predict_prob", 0.5),
            ))

        try:
            self.execute_many(insert_sql, records)
            logger.info("预测记录写入完成: %d 条", len(records))
            return len(records)
        except Exception as exc:
            logger.error("预测记录写入失败: %s", str(exc))
            return 0

    def update_backtest_results(self, backtest_df: pd.DataFrame) -> int:
        """
        回填回测结果（更新 actual_direction, actual_pct_change, is_correct）
        """
        if backtest_df is None or len(backtest_df) == 0:
            return 0

        update_sql = """
            UPDATE prediction_backtest_records
            SET actual_direction = %s,
                actual_pct_change = %s,
                is_correct = %s,
                backtest_time = %s
            WHERE predict_date = %s AND index_code = %s
        """

        records = []
        for _, row in backtest_df.iterrows():
            if pd.isna(row.get("is_correct")):
                continue
            records.append((
                str(row.get("actual_direction", "")),
                float(row.get("actual_pct_change", 0.0)),
                int(row.get("is_correct", 0)),
                row.get("backtest_time", datetime.now()),
                row.get("predict_date"),
                str(row.get("index_code", "")),
            ))

        if not records:
            return 0

        try:
            self.execute_many(update_sql, records)
            logger.info("回测结果回填完成: %d 条", len(records))
            return len(records)
        except Exception as exc:
            logger.error("回测结果回填失败: %s", str(exc))
            return 0

    def get_historical_sentiment(self, days: int = 60) -> pd.DataFrame:
        """获取历史情绪统计数据（供ML训练）"""
        sql = """
            SELECT stat_date, account_level, total_posts,
                   bullish_count, bearish_count, neutral_count,
                   bullish_ratio, bearish_ratio, neutral_ratio,
                   avg_sentiment_score
            FROM daily_sentiment_stats
            WHERE account_level = 'all'
            ORDER BY stat_date DESC
            LIMIT %s
        """
        return self.query(sql, (days,))

    def get_predictions_by_date(self, predict_date: date) -> pd.DataFrame:
        """获取指定日期的预测记录"""
        sql = """
            SELECT predict_date, target_date, index_code, index_name,
                   predict_direction, predict_prob,
                   actual_direction, actual_pct_change, is_correct, backtest_time
            FROM prediction_backtest_records
            WHERE predict_date = %s
            ORDER BY index_code
        """
        return self.query(sql, (predict_date,))

    def get_all_pending_backtest(self) -> pd.DataFrame:
        """获取所有待回测的预测记录 (is_correct IS NULL)"""
        sql = """
            SELECT predict_date, target_date, index_code, index_name,
                   predict_direction, predict_prob
            FROM prediction_backtest_records
            WHERE is_correct IS NULL
            ORDER BY predict_date, index_code
        """
        return self.query(sql)


# ==================== Excel报表导出 ====================

class ExcelReporter:
    """
    Excel日报导出器
    生成格式化的【日期_财经舆情预测日报.xlsx】
    """

    # 样式定义
    COLOR_GREEN = "92D050"   # 涨/正确
    COLOR_RED = "FF6B6B"     # 跌/错误
    COLOR_HEADER = "1F4E79"  # 表头深蓝
    COLOR_TITLE = "D6E4F0"   # 标题浅蓝
    COLOR_WHITE = "FFFFFF"
    COLOR_LIGHT_GRAY = "F2F2F2"

    def __init__(self):
        self.wb: Optional[Workbook] = None

    def generate_report(
        self,
        report_date: date,
        posts_df: pd.DataFrame,
        stats_df: pd.DataFrame,
        predictions: List[Dict],
        accuracy: Optional[Dict] = None,
    ) -> str:
        """
        生成完整的Excel日报

        Args:
            report_date: 报告日期
            posts_df: 帖子数据
            stats_df: 情绪统计数据
            predictions: 预测结果列表
            accuracy: 回测准确率数据（可选，仅在回测时有）

        Returns:
            导出文件路径
        """
        self.wb = Workbook()

        # Sheet 1: 情绪统计概览
        ws1 = self.wb.active
        ws1.title = "情绪统计概览"
        self._write_sentiment_sheet(ws1, report_date, stats_df)

        # Sheet 2: 涨跌预测结果
        ws2 = self.wb.create_sheet("涨跌预测结果")
        self._write_prediction_sheet(ws2, report_date, predictions, accuracy)

        # Sheet 3: 帖子明细（只保留前500条避免文件过大）
        ws3 = self.wb.create_sheet("帖子明细")
        if posts_df is not None and len(posts_df) > 0:
            sample_df = posts_df.head(500)
        else:
            sample_df = pd.DataFrame()
        self._write_posts_sheet(ws3, sample_df)

        # Sheet 4: 回测准确率趋势（如果有）
        if accuracy:
            ws4 = self.wb.create_sheet("回测准确率")
            self._write_accuracy_sheet(ws4, accuracy)

        # 保存
        file_path = EXCEL_REPORT_TEMPLATE.format(date=report_date.strftime("%Y%m%d"))
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        self.wb.save(file_path)
        logger.info("Excel日报已导出: %s", file_path)
        return file_path

    def _write_sentiment_sheet(self, ws, report_date: date, stats_df: pd.DataFrame):
        """写入情绪统计概览Sheet"""
        # 标题
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"财经舆情情绪统计日报 - {report_date.strftime('%Y-%m-%d')}"
        title_cell.font = Font(name="微软雅黑", size=16, bold=True, color=self.COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 表头
        headers = [
            "统计日期", "账号等级", "总帖子数",
            "看多数量", "看空数量", "中性数量",
            "看多占比", "看空占比", "中性占比",
        ]
        self._write_header(ws, headers, row=3)

        # 数据
        if stats_df is not None and len(stats_df) > 0:
            for i, (_, row) in enumerate(stats_df.iterrows(), start=4):
                ws.cell(row=i, column=1, value=str(row.get("stat_date", "")))
                ws.cell(row=i, column=2, value=str(row.get("account_level", "")))
                ws.cell(row=i, column=3, value=int(row.get("total_posts", 0)))
                ws.cell(row=i, column=4, value=int(row.get("bullish_count", 0)))
                ws.cell(row=i, column=5, value=int(row.get("bearish_count", 0)))
                ws.cell(row=i, column=6, value=int(row.get("neutral_count", 0)))
                self._write_pct(ws, i, 7, float(row.get("bullish_ratio", 0)))
                self._write_pct(ws, i, 8, float(row.get("bearish_ratio", 0)))
                self._write_pct(ws, i, 9, float(row.get("neutral_ratio", 0)))

        # 列宽
        self._auto_width(ws, len(headers))

    def _write_prediction_sheet(
        self, ws, report_date: date, predictions: List[Dict],
        accuracy: Optional[Dict] = None,
    ):
        """写入涨跌预测结果Sheet"""
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"四大指数次日涨跌预测(基于情感分析) - {report_date.strftime('%Y-%m-%d')}"
        title_cell.font = Font(name="微软雅黑", size=16, bold=True, color=self.COLOR_HEADER)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        headers = [
            "预测日期", "目标日期", "指数代码", "指数名称",
            "预测方向", "预测概率", "情感多空得分", "实际涨跌幅", "是否正确",
        ]
        self._write_header(ws, headers, row=3)

        if predictions:
            for i, pred in enumerate(predictions, start=4):
                ws.cell(row=i, column=1, value=str(pred.get("predict_date", "")))
                ws.cell(row=i, column=2, value=str(pred.get("target_date", "")))
                ws.cell(row=i, column=3, value=str(pred.get("index_code", "")))
                ws.cell(row=i, column=4, value=str(pred.get("index_name", "")))

                direction = str(pred.get("predict_direction", ""))
                cell_dir = ws.cell(row=i, column=5, value="看涨" if direction == "up" else "看跌")
                if direction == "up":
                    cell_dir.font = Font(color="CC0000", bold=True)
                else:
                    cell_dir.font = Font(color="006600", bold=True)

                ws.cell(row=i, column=6, value=float(pred.get("predict_prob", 0.5)))
                ws.cell(row=i, column=6).number_format = "0.00%"

                # 情感多空得分（综合分析指标）
                net_sentiment = pred.get("net_sentiment", 0)
                ws.cell(row=i, column=7, value=round(float(net_sentiment), 4))

                actual = pred.get("actual_pct_change")
                if actual is not None:
                    ws.cell(row=i, column=8, value=float(actual))
                    ws.cell(row=i, column=8).number_format = "+0.00%;-0.00%"

                correct = pred.get("is_correct")
                if correct is not None:
                    cell_correct = ws.cell(row=i, column=9,
                                           value="✓正确" if correct == 1 else "✗错误")
                    if correct == 1:
                        cell_correct.font = Font(color="006600", bold=True)
                    else:
                        cell_correct.font = Font(color="CC0000", bold=True)

        self._auto_width(ws, len(headers))

    def _write_posts_sheet(self, ws, posts_df: pd.DataFrame):
        """写入帖子明细Sheet"""
        ws.merge_cells("A1:H1")
        ws["A1"].value = "帖子明细（最近500条）"
        ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color=self.COLOR_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        if posts_df is None or len(posts_df) == 0:
            ws.cell(row=3, column=1, value="暂无数据")
            return

        headers = ["平台", "博主", "等级", "帖子类型", "情感", "情感得分", "点赞数", "正文摘要"]
        self._write_header(ws, headers, row=3)

        for i, (_, row) in enumerate(posts_df.iterrows(), start=4):
            ws.cell(row=i, column=1, value=str(row.get("platform", "")))
            ws.cell(row=i, column=2, value=str(row.get("account_name", "")))
            ws.cell(row=i, column=3, value=str(row.get("account_level", "")))
            ws.cell(row=i, column=4, value=str(row.get("post_type", "")))
            ws.cell(row=i, column=5, value=str(row.get("sentiment", "")))
            ws.cell(row=i, column=6, value=float(row.get("sentiment_score", 0.5)))
            ws.cell(row=i, column=7, value=int(row.get("like_count", 0)))
            content = str(row.get("post_content", ""))[:80]
            ws.cell(row=i, column=8, value=content)

        self._auto_width(ws, len(headers))

    def _write_accuracy_sheet(self, ws, accuracy: Dict):
        """写入回测准确率Sheet"""
        ws.merge_cells("A1:D1")
        ws["A1"].value = f"回测准确率统计 - {accuracy.get('backtest_date', '')}"
        ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color=self.COLOR_HEADER)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        headers = ["指标", "数值", "说明", ""]
        self._write_header(ws, headers, row=3)

        metrics = [
            ("回测日期", str(accuracy.get("backtest_date", "")), ""),
            ("总预测数", accuracy.get("total_predictions", 0), ""),
            ("正确数量", accuracy.get("correct_count", 0), ""),
            ("整体准确率", f"{accuracy.get('overall_accuracy', 0)*100:.2f}%", ""),
        ]
        for i, (label, value, note) in enumerate(metrics, start=4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=3, value=note)

        # 分指数
        per_index = accuracy.get("per_index_accuracy", {})
        if per_index:
            row_offset = len(metrics) + 5
            ws.cell(row=row_offset, column=1, value="分指数准确率").font = Font(bold=True)
            for j, (code, info) in enumerate(per_index.items()):
                r = row_offset + 1 + j
                ws.cell(row=r, column=1, value=f"{info['name']} ({code})")
                ws.cell(row=r, column=2,
                        value=f"{info['accuracy']*100:.2f}% ({info['correct']}/{info['total']})")

    # ==================== 样式辅助方法 ====================

    def _write_header(self, ws, headers: List[str], row: int):
        """写入统一表头样式"""
        header_fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER,
                                  fill_type="solid")
        header_font = Font(name="微软雅黑", size=11, bold=True, color=self.COLOR_WHITE)
        thin_border = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="thin", color="B0B0B0"),
            bottom=Side(style="thin", color="B0B0B0"),
        )
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[row].height = 25

    @staticmethod
    def _write_pct(ws, row: int, col: int, value: float):
        """写入百分比格式"""
        cell = ws.cell(row=row, column=col, value=value)
        cell.number_format = "0.00%"

    @staticmethod
    def _auto_width(ws, num_cols: int, min_width: int = 10, max_width: int = 25):
        """自适应列宽"""
        for col_idx in range(1, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_len = min_width
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                for cell_value in row:
                    if cell_value:
                        max_len = max(max_len, min(len(str(cell_value)) + 2, max_width))
            ws.column_dimensions[col_letter].width = max_len


# ==================== 便捷函数 ====================

def generate_excel_report(
    report_date: date,
    posts_df: pd.DataFrame,
    stats_df: pd.DataFrame,
    predictions: List[Dict],
    accuracy: Optional[Dict] = None,
) -> str:
    """便捷函数: 一键生成Excel日报"""
    reporter = ExcelReporter()
    return reporter.generate_report(report_date, posts_df, stats_df, predictions, accuracy)


# ==================== 本地测试入口 ====================
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")

    # 测试Excel导出（不需要数据库）
    import numpy as np
    test_posts = pd.DataFrame({
        "platform": ["douyin"] * 10,
        "account_name": [f"财经大V_{i}" for i in range(10)],
        "account_level": ["top1000"] * 5 + ["10w_plus"] * 5,
        "post_type": ["video"] * 5 + ["image_text"] * 5,
        "sentiment": np.random.choice(["bullish", "bearish", "neutral"], 10),
        "sentiment_score": np.random.uniform(0.3, 0.7, 10),
        "like_count": np.random.randint(0, 1000, 10),
        "post_content": ["测试内容" + str(i) * 20 for i in range(10)],
    })

    test_stats = pd.DataFrame({
        "stat_date": [date.today()] * 3,
        "account_level": ["all", "top1000", "10w_plus"],
        "total_posts": [500, 200, 300],
        "bullish_count": [220, 90, 130],
        "bearish_count": [150, 60, 90],
        "neutral_count": [130, 50, 80],
        "bullish_ratio": [0.44, 0.45, 0.4333],
        "bearish_ratio": [0.30, 0.30, 0.30],
        "neutral_ratio": [0.26, 0.25, 0.2667],
    })

    test_preds = [
        {"predict_date": date.today(), "target_date": date.today(),
         "index_code": code, "index_name": name,
         "predict_direction": "up", "predict_prob": 0.62,
         "net_sentiment": 0.12, "actual_pct_change": None, "is_correct": None}
        for code, name in INDEX_CODE_MAP.items()
    ]

    file_path = generate_excel_report(date.today(), test_posts, test_stats, test_preds)
    print(f"\nExcel测试报表已生成: {file_path}")